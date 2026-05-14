#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Basira Web Scraping Local
A lightweight, Python-based web scraping tool
No browser automation, no Node.js, no complex dependencies
"""

import os
import sys
import json
import time
import socket
import re
import html as html_lib
import mimetypes
from datetime import datetime
from urllib.parse import urljoin, urlparse, quote, unquote
from collections import Counter

from flask import Flask, render_template, request, jsonify, send_file, Response, make_response
from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ═══════════════════════════════════════════════════════════════════════════
# Configuration
# ═══════════════════════════════════════════════════════════════════════════

BASIRA_PORT = int(os.environ.get('BASIRA_PORT', 8797))
BASIRA_HOST = '127.0.0.1'

app = Flask(__name__)
app.config['JSON_AS_ASCII'] = False

# User-Agent for requests
USER_AGENT = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'


FRIENDLY_ERROR_MAP = [
    ('No module named', 'A required Python library is missing. Run: python install_libraries.py inside the project folder.'),
    ('403', 'The website rejected the request. Try reducing pages, checking the selector, or using a valid proxy.'),
    ('Access Denied', 'The website blocked the request. Try a valid proxy or choose a website that allows normal HTTP access.'),
    ('Connection', 'Basira could not connect to the website. Check the URL, internet connection, proxy settings, and whether the site is reachable.'),
]

def friendly_error(exc):
    """Convert technical exceptions into UI-safe, actionable messages."""
    text = str(exc)
    for needle, message in FRIENDLY_ERROR_MAP:
        if needle.lower() in text.lower():
            return message
    return text


def validate_url_value(url):
    if not url:
        return False, 'Website URL is required.'
    parsed = urlparse(url)
    if parsed.scheme not in ('http', 'https') or not parsed.netloc:
        return False, 'Enter a valid website URL, for example https://example.com.'
    return True, ''


def validate_scrape_options(max_rows, max_pages, max_load_more_clicks):
    if int(max_rows or 0) < 1:
        return False, 'Max Rows must be at least 1.'
    if int(max_rows or 0) > 50000:
        return False, 'Max Rows is too large. Use 50,000 rows or less.'
    if int(max_pages or 0) < 1:
        return False, 'Max Pages must be at least 1.'
    if int(max_pages or 0) > 200:
        return False, 'Max Pages is too large. Use 200 pages or less.'
    if int(max_load_more_clicks or 0) > 200:
        return False, 'Load More clicks is too large. Use 200 clicks or less.'
    return True, ''

# ═══════════════════════════════════════════════════════════════════════════
# Helper Functions
# ═══════════════════════════════════════════════════════════════════════════

def is_port_in_use(port):
    """Check if a port is already in use"""
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        return s.connect_ex(('127.0.0.1', port)) == 0

def clean_url(url):
    """Add https:// if not present"""
    if not url:
        return None
    url = url.strip()
    if not url.startswith(('http://', 'https://')):
        url = 'https://' + url
    return url

def build_proxy_config(proxy_url):
    """Build a requests-compatible proxy dictionary from a user proxy URL."""
    proxy_url = (proxy_url or '').strip()
    if not proxy_url:
        return None
    if not re.match(r'^[a-zA-Z][a-zA-Z0-9+.-]*://', proxy_url):
        proxy_url = 'http://' + proxy_url
    return {'http': proxy_url, 'https': proxy_url}


def fetch_page_requests(url, timeout=15, proxy_url=''):
    """Fetch a web page with requests, optionally through a proxy."""
    headers = {
        'User-Agent': USER_AGENT,
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        'Accept-Language': 'en-US,en;q=0.9,ar;q=0.8',
    }
    proxies = build_proxy_config(proxy_url)
    try:
        response = requests.get(url, headers=headers, timeout=timeout, proxies=proxies)
        response.raise_for_status()
        if not response.encoding:
            response.encoding = response.apparent_encoding or 'utf-8'
        return response.text
    except requests.exceptions.RequestException as e:
        raise Exception(friendly_error(e))


def fetch_page_browser(*args, **kwargs):
    """Browser rendering is intentionally disabled in the NO-CHROMIUM edition."""
    raise Exception(
        'This Basira edition is requests-only and does not use Chromium or Playwright. Turn off Render JavaScript and Load More, then scrape again.'
    )


def fetch_page(url, timeout=15, proxy_url='', render=False, load_more_selector='', max_load_more_clicks=0, wait_after_load=1200):
    """Fetch a web page with requests only. Chromium and Playwright are intentionally not used."""
    if render or (load_more_selector and int(max_load_more_clicks or 0) > 0):
        raise Exception(
            'This Basira edition is requests-only and does not use Chromium or Playwright. Turn off Render JavaScript and Load More, then scrape again.'
        )
    return fetch_page_requests(url, timeout=timeout, proxy_url=proxy_url)





def rewrite_css_urls(css_text, css_url, proxy_url=''):
    """Rewrite CSS url(...) and @import assets through Basira so the visual preview keeps its real styling."""
    if not css_text:
        return ''

    def repl(match):
        raw = match.group(1).strip().strip('"\'')
        if not raw or raw.startswith(('data:', 'javascript:', '#')):
            return match.group(0)
        absolute = urljoin(css_url, raw)
        proxied = proxify_asset_url(absolute, proxy_url)
        return f'url("{proxied}")'

    css_text = re.sub(r'url\(([^)]+)\)', repl, css_text, flags=re.IGNORECASE)

    def import_repl(match):
        quote_char = match.group(1) or ''
        raw = (match.group(2) or '').strip()
        if not raw or raw.startswith(('data:', 'javascript:', '#')):
            return match.group(0)
        absolute = urljoin(css_url, raw)
        return f'@import url("{proxify_asset_url(absolute, proxy_url)}")'

    css_text = re.sub(r'@import\s+(?:url\()?([\"\']?)([^\"\')\s;]+)\1\)?', import_repl, css_text, flags=re.IGNORECASE)
    return css_text


def proxify_asset_url(asset_url, proxy_url=''):
    """Build an absolute Basira URL for a remote visual-preview asset.
    It must be absolute because the preview page uses a <base> tag for the target site.
    """
    if not asset_url:
        return ''
    try:
        local_base = request.host_url.rstrip('/')
    except Exception:
        local_base = ''
    proxied = local_base + '/visual-asset?url=' + quote(asset_url, safe='')
    if proxy_url:
        proxied += '&proxy=' + quote(proxy_url, safe='')
    return proxied



def fully_unquote_url(value, max_rounds=3):
    """Decode nested percent-encoding used by CSS font URLs, for example %253F -> ?."""
    current = (value or '').strip()
    for _ in range(max_rounds):
        decoded = unquote(current)
        if decoded == current:
            break
        current = decoded
    return current

def empty_asset_response(asset_url):
    """Return a harmless empty response for optional assets such as web fonts.
    This keeps the visual selector styled even when a remote font file is blocked.
    """
    guessed = mimetypes.guess_type(asset_url.split('?', 1)[0])[0] or 'application/octet-stream'
    resp = Response(b'', content_type=guessed)
    resp.headers['Access-Control-Allow-Origin'] = '*'
    resp.headers['Cache-Control'] = 'public, max-age=3600'
    return resp

def build_visual_proxy_html(target_url, timeout=15, proxy_url=''):
    """Fetch a target page and return a same-origin, static HTML preview for visual selection.
    This avoids iframe/X-Frame-Options/CSP blocking because the iframe loads from Basira itself.
    """
    headers = {
        'User-Agent': USER_AGENT,
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        'Accept-Language': 'en-US,en;q=0.9,ar;q=0.8'
    }
    response = requests.get(target_url, headers=headers, timeout=timeout, proxies=build_proxy_config(proxy_url))
    response.raise_for_status()

    if not response.encoding:
        response.encoding = response.apparent_encoding or 'utf-8'

    soup = BeautifulSoup(response.text, 'html.parser')

    for tag in soup.find_all('script'):
        tag.decompose()

    for meta in soup.find_all('meta'):
        http_equiv = (meta.get('http-equiv') or '').lower()
        if http_equiv in ('content-security-policy', 'refresh'):
            meta.decompose()

    if soup.head is None:
        head = soup.new_tag('head')
        if soup.html:
            soup.html.insert(0, head)
        else:
            soup.insert(0, head)
    else:
        head = soup.head

    for old_base in head.find_all('base'):
        old_base.decompose()
    base = soup.new_tag('base', href=target_url)
    head.insert(0, base)
    meta_referrer = soup.new_tag('meta', attrs={'name': 'referrer', 'content': 'no-referrer-when-downgrade'})
    head.insert(1, meta_referrer)

    # Normalize protocol-relative and http assets before proxying them through Basira.
    for tag, attrs in {
        'link': ['href'],
        'img': ['src', 'data-src', 'data-original', 'data-lazy'],
        'source': ['src', 'srcset'],
        'script': ['src'],
        'a': ['href']
    }.items():
        for element in soup.find_all(tag):
            for attr in attrs:
                value = element.get(attr)
                if isinstance(value, str) and value.startswith('//'):
                    element[attr] = 'https:' + value

    # Keep all preview assets same-origin. This prevents browser CORS noise for fonts, CSS, and images
    # while still rendering the page inside Basira's visual selector.
    for link in soup.find_all('link'):
        href = link.get('href')
        rel = ' '.join(link.get('rel') or []).lower()
        if href and ('stylesheet' in rel or href.lower().endswith('.css')):
            link['href'] = proxify_asset_url(urljoin(target_url, href), proxy_url)
            for noisy_attr in ['integrity', 'crossorigin']:
                if link.has_attr(noisy_attr):
                    del link[noisy_attr]

    for tag_name, attr_names in {
        'img': ['src', 'data-src', 'data-original', 'data-lazy'],
        'source': ['src', 'srcset']
    }.items():
        for element in soup.find_all(tag_name):
            for attr in attr_names:
                value = element.get(attr)
                if isinstance(value, str) and value.strip():
                    if attr == 'srcset':
                        parts = []
                        for part in value.split(','):
                            tokens = part.strip().split()
                            if not tokens:
                                continue
                            tokens[0] = proxify_asset_url(urljoin(target_url, tokens[0]), proxy_url)
                            parts.append(' '.join(tokens))
                        element[attr] = ', '.join(parts)
                    else:
                        element[attr] = proxify_asset_url(urljoin(target_url, value), proxy_url)

    # Also proxy common media/poster attributes and inline CSS background images.
    for element in soup.find_all(True):
        for attr in ['poster', 'data-srcset']:
            value = element.get(attr)
            if isinstance(value, str) and value.strip():
                if attr.endswith('srcset'):
                    parts = []
                    for part in value.split(','):
                        tokens = part.strip().split()
                        if tokens:
                            tokens[0] = proxify_asset_url(urljoin(target_url, tokens[0]), proxy_url)
                            parts.append(' '.join(tokens))
                    element[attr] = ', '.join(parts)
                else:
                    element[attr] = proxify_asset_url(urljoin(target_url, value), proxy_url)
        style_value = element.get('style')
        if isinstance(style_value, str) and 'url(' in style_value.lower():
            element['style'] = rewrite_css_urls(style_value, target_url, proxy_url)


    for style_tag in soup.find_all('style'):
        if style_tag.string:
            style_tag.string.replace_with(rewrite_css_urls(str(style_tag.string), target_url, proxy_url))

    # Prefer actual lazy image URLs in the preview.
    for img in soup.find_all('img'):
        for lazy_attr in ['data-src', 'data-original', 'data-lazy']:
            if img.get(lazy_attr):
                img['src'] = img.get(lazy_attr)
                break

    for a in soup.find_all('a'):
        href = a.get('href')
        if href:
            a['data-basira-href'] = urljoin(target_url, href)
            a['href'] = 'javascript:void(0)'


    # Reduce accessibility warnings in the local preview by ensuring form controls have names/labels.
    for idx, control in enumerate(soup.find_all(['input', 'select', 'textarea']), 1):
        if not control.get('id'):
            control['id'] = f'basira-preview-field-{idx}'
        if not control.get('name'):
            control['name'] = control.get('id')
        if not control.get('aria-label'):
            control['aria-label'] = control.get('placeholder') or control.get('name') or 'form field'

    return str(soup)



def normalize_text(value):
    """Clean common encoding artifacts and whitespace."""
    if value is None:
        return ''
    text = str(value)
    replacements = {
        'Â£': '£',
        'Â€': '€',
        'Â¥': '¥',
        'Â©': '©',
        'Â®': '®',
        'Â ': ' ',
        'Â': '',
        '\u00a0': ' ',
        '\ufeff': '',
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def css_escape(value):
    """Minimal CSS identifier escaping for generated selectors."""
    value = str(value or '')
    return re.sub(r'([^a-zA-Z0-9_-])', lambda m: '\\' + m.group(1), value)


def class_is_stable(class_name):
    """Avoid layout/state classes while keeping meaningful names such as price_color or product_pod."""
    c = (class_name or '').lower().strip()
    if not c:
        return False
    exact_noisy = {
        'active', 'selected', 'hover', 'focus', 'hidden', 'show', 'open', 'close',
        'row', 'col', 'grid', 'container', 'wrapper', 'clearfix', 'flex', 'button', 'btn'
    }
    prefix_noisy = (
        'col-', 'row-', 'grid-', 'text-', 'bg-', 'mt-', 'mb-', 'ml-', 'mr-',
        'pt-', 'pb-', 'pl-', 'pr-', 'px-', 'py-', 'justify-', 'items-', 'w-', 'h-'
    )
    if c in exact_noisy:
        return False
    if any(c.startswith(prefix) for prefix in prefix_noisy):
        return False
    # Keep content/data classes, for example price_color, product_pod, availability, thumbnail.
    return True


def best_selector_for_element(el, root_soup=None):
    """Generate a practical selector for an element."""
    if not el or not getattr(el, 'name', None):
        return ''
    if el.get('id'):
        candidate = f"#{css_escape(el.get('id'))}"
        if not root_soup or len(root_soup.select(candidate)) == 1:
            return candidate
    classes = [c for c in (el.get('class') or []) if class_is_stable(c)]
    if classes:
        joined = ''.join('.' + css_escape(c) for c in classes[:2])
        candidate = f"{el.name}{joined}"
        if root_soup and len(root_soup.select(candidate)) >= 1:
            return candidate
        return joined
    return el.name


def relative_selector(child, parent):
    """Generate a short selector to find child inside item parent."""
    if not child or child is parent:
        return ''
    parts = []
    node = child
    while node and node is not parent and getattr(node, 'name', None):
        part = best_selector_for_element(node)
        if not part:
            part = node.name
        parts.insert(0, part)
        if len(parts) >= 3:
            break
        node = node.parent
    # Prefer simple meaningful selector when it works from parent.
    simple_options = []
    if child.get('class'):
        simple_options.append(best_selector_for_element(child))
    if child.name in ['h1', 'h2', 'h3', 'h4', 'h5', 'h6']:
        simple_options.append(child.name)
    if child.name == 'a':
        heading_parent = child.find_parent(['h1', 'h2', 'h3', 'h4', 'h5', 'h6'])
        if heading_parent and heading_parent is not parent:
            simple_options.append(f'{heading_parent.name} a')
        simple_options.append('a')
    if child.name == 'img':
        simple_options.append('img')
    for opt in simple_options:
        try:
            if opt and parent.select_one(opt) is child:
                return opt
        except Exception:
            pass
    return ' '.join(parts)


def detect_next_selector(soup):
    """Suggest a next-page selector."""
    candidates = [
        'a[rel="next"]', 'link[rel="next"]', '.next a', 'li.next a',
        '.pagination .next a', '.pagination a.next', 'a.next', 'a.next-page',
        '.pager a.next', 'a[aria-label*="Next" i]', 'a[title*="Next" i]'
    ]
    for sel in candidates:
        try:
            if soup.select_one(sel):
                return sel
        except Exception:
            continue
    for a in soup.find_all('a', href=True):
        text = normalize_text(a.get_text(' ', strip=True)).lower()
        if text in ['next', 'next page', '>', '›', '»'] or 'next' in text:
            return relative_selector(a, soup) or 'a'
    return ''


def infer_field_name(selector, field_type, sample):
    selector_l = (selector or '').lower()
    sample_l = (sample or '').lower()
    if field_type == 'image':
        return 'Image'
    if field_type == 'link':
        return 'Link'
    if 'price' in selector_l or re.search(r'[£$€¥]|\bSAR\b|ريال', sample, re.I):
        return 'Price'
    if 'stock' in sample_l or 'available' in sample_l or 'availability' in selector_l:
        return 'Availability'
    if 'rating' in selector_l or 'star' in selector_l:
        return 'Rating'
    if any(h in selector_l for h in ['h1', 'h2', 'h3', 'title', 'name']):
        return 'Title'
    return 'Field'


def make_unique_field_names(fields):
    seen = {}
    for f in fields:
        base = re.sub(r'[^A-Za-z0-9_]+', '_', f.get('name') or 'Field').strip('_') or 'Field'
        count = seen.get(base, 0) + 1
        seen[base] = count
        f['name'] = base if count == 1 else f'{base}_{count}'
    return fields


def classify_detected_group(selector, fields):
    selector_l = (selector or '').lower()
    names = {str(f.get('name', '')).lower() for f in (fields or [])}
    if any(x in selector_l for x in ['product', 'item', 'card', 'listing', 'result', 'pod']) or {'title', 'price'}.issubset(names):
        return 'Products'
    if any(x in selector_l for x in ['article', 'post', 'entry', 'blog']):
        return 'Articles'
    if 'link' in names and len(names) <= 2:
        return 'Links'
    return 'Repeated items'


def compact_selector_label(selector):
    selector = (selector or '').strip()
    if not selector:
        return 'Auto'
    # Prefer a concrete selector in the UI. Broad contains selectors are valid but less readable.
    selector = selector.replace('[class*="product" i]', 'product-like elements')
    selector = selector.replace('[class*="card" i]', 'card-like elements')
    selector = selector.replace('[class*="item" i]', 'item-like elements')
    return selector


def refine_item_selector(selector, elements):
    """Replace broad class contains selectors with the most concrete repeated card selector available."""
    selector = (selector or '').strip()
    if not elements:
        return selector
    broad = bool(re.search(r'\[class\*=', selector, re.I)) or selector in ('article', 'li', 'div')
    if not broad:
        return selector
    common = None
    for el in elements[:20]:
        classes = set(el.get('class') or [])
        classes = {c for c in classes if class_is_stable(c)}
        common = classes if common is None else common.intersection(classes)
    if common:
        priority_words = ('product_pod', 'product', 'card', 'item', 'listing', 'result', 'article', 'post', 'entry')
        ordered = sorted(common, key=lambda c: (0 if any(w in c.lower() for w in priority_words) else 1, len(c)))
        cls = ordered[0]
        if elements[0].name and elements[0].name not in ('div', 'span'):
            return f"{elements[0].name}.{css_escape(cls)}"
        return f".{css_escape(cls)}"
    return selector


def detect_fields_for_item(item, base_url):
    """Infer useful fields from the first repeated item."""
    fields = []
    used = set()

    def add(name, selector, field_type, sample=''):
        selector = (selector or '').strip()
        key = (selector, field_type)
        if not selector or key in used:
            return
        used.add(key)
        fields.append({'name': name, 'selector': selector, 'type': field_type, 'sample': normalize_text(sample)[:120]})

    # Title-like text. Prefer links inside headings so titles are not mixed with image links.
    title = None
    for title_selector in [
        'h1 a[title]', 'h2 a[title]', 'h3 a[title]', 'h4 a[title]',
        'h1 a', 'h2 a', 'h3 a', 'h4 a',
        '[class*="title" i] a', '[class*="name" i] a',
        'h1', 'h2', 'h3', 'h4', '[class*="title" i]', '[class*="name" i]'
    ]:
        try:
            title = item.select_one(title_selector)
        except Exception:
            title = None
        if title:
            break
    if title:
        add('Title', relative_selector(title, item), 'text', extract_text(title))

    # Price-like text. Prefer known price selectors and the smallest element that contains a currency value.
    price = None
    for price_selector in ['.price_color', '.price', '[class*=\"price\" i]', '[itemprop*=\"price\" i]', '[data-price]']:
        try:
            price = item.select_one(price_selector)
        except Exception:
            price = None
        if price:
            break
    currency_re = re.compile(r'([£$€¥]|\bSAR\b|ريال)\s*\d|\d+[\.,]?\d*\s*(SAR|ريال)', re.I)
    for el in item.find_all(['span', 'p', 'strong', 'b', 'em', 'div']):
        txt = normalize_text(el.get_text(' ', strip=True))
        if currency_re.search(txt):
            nested_currency = False
            for child in el.find_all(['span', 'p', 'strong', 'b', 'em', 'div']):
                if child is not el and currency_re.search(normalize_text(child.get_text(' ', strip=True))):
                    nested_currency = True
                    break
            if not nested_currency:
                price = el
                break
    if price:
        add('Price', relative_selector(price, item), 'price', extract_text(price))

    # Link. Prefer title/detail links over image links.
    link = None
    for link_selector in ['h1 a[href]', 'h2 a[href]', 'h3 a[href]', 'h4 a[href]', 'a[title][href]', 'a[href]']:
        link = item.select_one(link_selector)
        if link:
            break
    if link:
        add('Link', relative_selector(link, item), 'link', extract_link(link, base_url))

    # Image.
    img = item.select_one('img')
    if img:
        add('Image', relative_selector(img, item), 'image', extract_image(img, base_url))

    # Availability/status.
    avail = item.select_one('[class*="avail" i], [class*="stock" i], [class*="status" i]')
    if not avail:
        for el in item.find_all(['p', 'span', 'div']):
            txt = normalize_text(el.get_text(' ', strip=True)).lower()
            if any(word in txt for word in ['in stock', 'out of stock', 'available', 'unavailable', 'متوفر', 'غير متوفر']):
                avail = el
                break
    if avail:
        add('Availability', relative_selector(avail, item), 'text', extract_text(avail))

    # Rating.
    rating = item.select_one('[class*="rating" i], [class*="star" i], [aria-label*="rating" i]')
    if rating:
        sample = ' '.join(rating.get('class') or []) or rating.get('aria-label') or extract_text(rating)
        selector = '[class*="rating" i]'
        try:
            if item.select_one(selector) is not rating:
                selector = relative_selector(rating, item)
        except Exception:
            selector = relative_selector(rating, item)
        add('Rating', selector, 'text', sample)

    # Fallback text fields if few fields detected.
    if len(fields) < 2:
        for el in item.find_all(['h1', 'h2', 'h3', 'h4', 'p', 'span', 'a']):
            txt = extract_text(el)
            if len(txt) >= 3:
                sel = relative_selector(el, item)
                add(infer_field_name(sel, 'text', txt), sel, 'text', txt)
                if len(fields) >= 5:
                    break

    return make_unique_field_names(fields[:8])


def detect_repeated_items(soup, base_url):
    """Find repeated product/card/list-item blocks and infer fields."""
    selector_counts = {}
    element_by_selector = {}
    candidates = []

    for el in soup.find_all(True):
        if el.name in ['html', 'head', 'body', 'script', 'style', 'meta', 'link', 'svg', 'path', 'nav', 'footer', 'header', 'form']:
            continue
        text = normalize_text(el.get_text(' ', strip=True))
        if len(text) < 8:
            continue
        selector = best_selector_for_element(el, soup)
        if not selector:
            continue
        try:
            count = len(soup.select(selector))
        except Exception:
            continue
        if count < 2 or count > 300:
            continue
        selector_counts[selector] = count
        element_by_selector.setdefault(selector, el)

    for selector, count in selector_counts.items():
        elements = soup.select(selector)
        sample_elements = elements[:min(5, len(elements))]
        avg_text = sum(len(normalize_text(e.get_text(' ', strip=True))) for e in sample_elements) / max(1, len(sample_elements))
        avg_children = sum(len(e.find_all(True)) for e in sample_elements) / max(1, len(sample_elements))
        with_links = sum(1 for e in sample_elements if e.select_one('a[href]'))
        with_images = sum(1 for e in sample_elements if e.select_one('img'))
        inferred_fields = detect_fields_for_item(elements[0], base_url)
        field_names = {f.get('name') for f in inferred_fields}
        class_bonus = 10 if selector.startswith('.') or '.' in selector else 0
        semantic_bonus = 0
        if re.search(r'product|item|card|post|article|result|listing|entry|book|pod', selector, re.I):
            semantic_bonus += 38
        if re.search(r'price|title|name|image|rating|stock|button|btn', selector, re.I):
            semantic_bonus -= 28
        if re.search(r'\[class\*=', selector, re.I):
            semantic_bonus -= 35
        if selector in ('article.product_pod', '.product_pod'):
            semantic_bonus += 140
        tag_penalty = 0
        if selector in ['a', 'p', 'span', 'li', 'div', 'h1', 'h2', 'h3', 'h4']:
            tag_penalty += 45
        if avg_children < 2:
            tag_penalty += 30
        field_bonus = len(inferred_fields) * 14
        if {'Title', 'Price'}.issubset(field_names):
            field_bonus += 35
        if 'Image' in field_names:
            field_bonus += 12
        if 'Link' in field_names:
            field_bonus += 12
        score = (min(count, 60) * 1.2) + min(avg_text, 320) / 8 + min(avg_children, 35) * 1.4 + with_links * 10 + with_images * 10 + class_bonus + semantic_bonus + field_bonus - tag_penalty
        if avg_text > 2500:
            score -= 60
        if len(inferred_fields) < 2:
            score -= 50
        candidates.append({
            'selector': refine_item_selector(selector, elements),
            'count': count,
            'score': round(score, 2),
            'sample': normalize_text(elements[0].get_text(' ', strip=True))[:180],
            'fields': inferred_fields
        })


    # Explicit product/list-card patterns. This makes Auto Detect choose real cards instead of inner text nodes.
    semantic_selectors = [
        'article.product_pod', '.product_pod', 'li.product', '.product',
        '[class*="product" i]', '[class*="card" i]', '[class*="item" i]',
        '[class*="listing" i]', '[class*="result" i]', 'article', 'li'
    ]
    seen_selectors = {c['selector'] for c in candidates}
    for selector in semantic_selectors:
        if selector in seen_selectors:
            continue
        try:
            elements = soup.select(selector)
        except Exception:
            continue
        elements = [e for e in elements if e.name not in ['html', 'head', 'body']]
        if len(elements) < 2 or len(elements) > 300:
            continue
        sample_elements = elements[:min(5, len(elements))]
        avg_text = sum(len(normalize_text(e.get_text(' ', strip=True))) for e in sample_elements) / max(1, len(sample_elements))
        avg_children = sum(len(e.find_all(True)) for e in sample_elements) / max(1, len(sample_elements))
        if avg_text < 8 or avg_children < 2:
            continue
        fields = detect_fields_for_item(elements[0], base_url)
        if len(fields) < 2:
            continue
        field_names = {f.get('name') for f in fields}
        score = 120 + min(len(elements), 60) + len(fields) * 18 + min(avg_children, 35) * 1.2
        if {'Title', 'Price'}.issubset(field_names):
            score += 45
        if 'Image' in field_names:
            score += 20
        if 'Link' in field_names:
            score += 20
        candidates.append({
            'selector': refine_item_selector(selector, elements),
            'count': len(elements),
            'score': round(score, 2),
            'sample': normalize_text(elements[0].get_text(' ', strip=True))[:180],
            'fields': fields
        })
        seen_selectors.add(selector)

    
    def _candidate_rank(c):
        selector = c.get('selector') or ''
        concrete_bonus = 0
        if selector == 'article.product_pod':
            concrete_bonus += 1000
        if '[class*=' in selector:
            concrete_bonus -= 500
        if selector.startswith('.') or '.' in selector:
            concrete_bonus += 80
        if selector in ('article', 'li', 'div', 'a', 'p', 'span'):
            concrete_bonus -= 250
        return (c.get('score', 0) + concrete_bonus, c.get('count', 0))
    candidates.sort(key=_candidate_rank, reverse=True)
    return candidates[:10]


def auto_detect_config(url, proxy_url=""):
    """Build an automatic scraping configuration for a URL."""
    html = fetch_page(url, proxy_url=proxy_url)
    soup = BeautifulSoup(html, 'html.parser')
    candidates = detect_repeated_items(soup, url)
    if candidates:
        best = candidates[0]
        fields = best.get('fields') or []
        if not fields:
            fields = [{'name': 'Text', 'selector': '*', 'type': 'text'}]
        clean_fields = [{k: v for k, v in f.items() if k != 'sample'} for f in fields]
        group = classify_detected_group(best['selector'], clean_fields)
        field_names = [f.get('name', 'Field') for f in clean_fields]
        return {
            'success': True,
            'detectedGroup': group,
            'itemSelector': best['selector'],
            'itemSelectorLabel': compact_selector_label(best['selector']),
            'itemCount': best.get('count', 0),
            'containerSelector': '',
            'fields': clean_fields,
            'fieldSamples': fields,
            'fieldNames': field_names,
            'paginationSelector': detect_next_selector(soup),
            'candidates': candidates,
            'summary': f"Detected: {group}. Item selector: {compact_selector_label(best['selector'])}. {best.get('count', 0)} items found. Fields detected: {', '.join(field_names) or 'None'}"
        }
    return {
        'success': False,
        'error': 'Could not detect a repeated product/list pattern automatically.'
    }

def extract_text(element):
    """Extract visible text. For title links, prefer the full title/alt attribute over truncated display text."""
    if not element:
        return ''
    attr_text = ''
    if getattr(element, 'name', '') in ('a', 'img'):
        attr_text = element.get('title') or element.get('alt') or ''
    text = normalize_text(element.get_text(' ', strip=True))
    attr_text = normalize_text(attr_text)
    if attr_text and (not text or '...' in text or len(attr_text) > len(text)):
        return attr_text
    return text

def extract_link(element, base_url):
    """Extract and convert relative link to absolute"""
    if not element:
        return ''
    href = element.get('href', '')
    if href:
        return urljoin(base_url, href)
    return ''

def extract_image(element, base_url):
    """Extract image src, checking multiple attributes"""
    if not element:
        return ''
    # Check common image source attributes
    for attr in ['src', 'data-src', 'data-lazy', 'data-original']:
        img_src = element.get(attr, '')
        if img_src:
            return urljoin(base_url, img_src)
    return ''

def extract_price(element):
    """Extract and clean price text"""
    if not element:
        return ''
    text = extract_text(element)
    # Remove common currency symbols and clean up
    import re
    # Keep numbers, dots, commas
    cleaned = re.sub(r'[^\d.,]+', '', text)
    return cleaned.strip()

def get_link_type(url, base_domain):
    """Determine if link is internal or external"""
    try:
        parsed = urlparse(url)
        link_domain = parsed.netloc
        if not link_domain:
            return 'internal'
        if base_domain in link_domain or link_domain in base_domain:
            return 'internal'
        return 'external'
    except:
        return 'unknown'

# ═══════════════════════════════════════════════════════════════════════════
# Routes
# ═══════════════════════════════════════════════════════════════════════════



@app.route('/visual-asset')
def visual_asset():
    """Proxy CSS, fonts, and images for the local visual preview.
    This avoids CORS errors caused by pages that reference remote fonts from a local preview origin.
    """
    try:
        raw_url = request.args.get('url', '')
        asset_url = fully_unquote_url(raw_url)
        proxy_url = fully_unquote_url(request.args.get('proxy', ''))
        # Some CSS files encode the query delimiter twice, producing font.woff%3Fv=...
        # Requests treats that as part of the path and the remote server may return 404.
        asset_url = asset_url.replace('%3F', '?').replace('%3f', '?')
        parsed = urlparse(asset_url)
        if parsed.scheme not in ('http', 'https'):
            return Response('Invalid asset URL', status=400, mimetype='text/plain')

        headers = {
            'User-Agent': USER_AGENT,
            'Accept': '*/*',
            'Referer': f'{parsed.scheme}://{parsed.netloc}/'
        }
        r = requests.get(asset_url, headers=headers, timeout=20, allow_redirects=True, proxies=build_proxy_config(proxy_url))
        r.raise_for_status()

        content_type = r.headers.get('Content-Type', '').split(';', 1)[0] or mimetypes.guess_type(asset_url.split('?', 1)[0])[0] or 'application/octet-stream'
        body = r.content

        if 'text/css' in content_type or asset_url.lower().split('?', 1)[0].endswith('.css'):
            encoding = r.encoding or r.apparent_encoding or 'utf-8'
            css = r.content.decode(encoding, errors='replace')
            css = rewrite_css_urls(css, asset_url, proxy_url)
            resp = Response(css, content_type='text/css; charset=utf-8')
        else:
            resp = Response(body, content_type=content_type)

        resp.headers['Access-Control-Allow-Origin'] = '*'
        resp.headers['Cache-Control'] = 'public, max-age=3600'
        return resp
    except Exception as e:
        failed_url = fully_unquote_url(request.args.get('url', '')).replace('%3F', '?').replace('%3f', '?')
        if re.search(r'\.(woff2?|ttf|eot|otf)(\?|$)', failed_url, re.I):
            return empty_asset_response(failed_url)
        return Response(f'Asset could not be loaded: {e}', status=502, mimetype='text/plain; charset=utf-8')


@app.route('/favicon.ico')
def favicon():
    return ('', 204)


@app.route('/visual-proxy')
def visual_proxy():
    """Same-origin static preview for the visual selector."""
    try:
        url = clean_url(request.args.get('url', ''))
        if not url:
            return Response('<h2>URL is required</h2>', status=400, mimetype='text/html; charset=utf-8')

        proxy_url = unquote(request.args.get('proxy', '')).strip()
        html = build_visual_proxy_html(url, proxy_url=proxy_url)
        return Response(html, mimetype='text/html; charset=utf-8')
    except Exception as e:
        safe_error = str(e).replace('<', '&lt;').replace('>', '&gt;')
        html = f"""<!doctype html><html><head><meta charset='utf-8'><style>
        body{{font-family:system-ui;background:#0f172a;color:#f8fafc;display:flex;min-height:100vh;align-items:center;justify-content:center;margin:0}}
        .box{{max-width:680px;padding:28px;border:1px solid #ef4444;border-radius:18px;background:#111827}}
        h2{{color:#ef4444;margin-top:0}}
        code{{color:#38bdf8;word-break:break-word}}
        </style></head><body><div class='box'><h2>Basira could not load this page</h2><p>{safe_error}</p><p>Try another URL or use Analyze Page / Test Selector.</p></div></body></html>"""
        return Response(html, status=502, mimetype='text/html; charset=utf-8')

@app.route('/')
def index():
    """Render main Basira interface"""
    resp = make_response(render_template('index.html'))
    resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    resp.headers['Pragma'] = 'no-cache'
    return resp


@app.route('/api/auto-detect', methods=['POST'])
def auto_detect():
    """Automatically detect the repeated item selector, useful fields, and pagination selector."""
    try:
        data = request.json or {}
        url = clean_url(data.get('url'))
        ok, message = validate_url_value(url)
        if not ok:
            return jsonify({'success': False, 'error': message}), 400
        proxy_enabled = bool(data.get('proxyEnabled', False))
        proxy_url = data.get('proxyUrl', '') if proxy_enabled else ''
        result = auto_detect_config(url, proxy_url=proxy_url)
        status = 200 if result.get('success') else 400
        return jsonify(result), status
    except Exception as e:
        return jsonify({'success': False, 'error': friendly_error(e)}), 500

@app.route('/api/scrape', methods=['POST'])
def scrape():
    """Execute structured data scraping with optional pagination."""
    try:
        data = request.json or {}
        url = clean_url(data.get('url'))
        container_selector = data.get('containerSelector', '').strip()
        item_selector = data.get('itemSelector', '').strip()
        fields = data.get('fields', [])
        max_rows = int(data.get('maxRows') or 100)
        max_pages = int(data.get('maxPages') or 1)
        pagination_selector = data.get('paginationSelector', '').strip()
        scroll_enabled = bool(data.get('scrollEnabled', False)) or max_pages > 1 or bool(pagination_selector)
        proxy_enabled = bool(data.get('proxyEnabled', False))
        proxy_url = data.get('proxyUrl', '').strip() if proxy_enabled else ''
        # NO-CHROMIUM edition: browser-only options are ignored intentionally.
        load_more_enabled = False
        load_more_selector = ''
        max_load_more_clicks = 0
        wait_after_load = int(data.get('waitAfterLoad') or 1200)
        render_enabled = False

        ok, message = validate_url_value(url)
        if not ok:
            return jsonify({'success': False, 'error': message}), 400
        ok, message = validate_scrape_options(max_rows, max_pages, max_load_more_clicks)
        if not ok:
            return jsonify({'success': False, 'error': message}), 400

        fields = [f for f in fields if f.get('name') and f.get('selector')]

        # Automatic mode: when the user only provides a URL, Basira detects the repeated item block and fields.
        auto_config_used = None
        if not item_selector or not fields:
            detected = auto_detect_config(url, proxy_url=proxy_url)
            if not detected.get('success'):
                return jsonify({'success': False, 'error': detected.get('error', 'Auto detection failed')}), 400
            if not item_selector:
                item_selector = detected.get('itemSelector', '')
            if not container_selector:
                container_selector = detected.get('containerSelector', '')
            if not fields:
                fields = detected.get('fields', [])
            if not pagination_selector:
                pagination_selector = detected.get('paginationSelector', '')
            auto_config_used = detected

        if not item_selector:
            return jsonify({'success': False, 'error': 'Item selector could not be detected'}), 400
        if not fields:
            return jsonify({'success': False, 'error': 'Fields could not be detected'}), 400

        max_rows = max(1, max_rows)
        max_pages = max(1, max_pages)
        pages_limit = max_pages if scroll_enabled else 1

        results = []
        failed_items = []
        visited_urls = []
        current_url = url

        common_next_selectors = [
            'a[rel="next"]',
            'link[rel="next"]',
            '.next a',
            'li.next a',
            '.pagination .next a',
            '.pagination a.next',
            'a.next',
            'a.next-page',
            '.pager a.next',
            'nav.pagination a[aria-label*="Next" i]',
            'a[aria-label*="Next" i]',
        ]

        def scoped_soup(page_soup):
            if not container_selector:
                return page_soup
            container = page_soup.select_one(container_selector)
            return container if container else None

        def extract_row(item, base_url):
            row = {}
            for field in fields:
                field_name = field.get('name', '').strip()
                field_selector = field.get('selector', '').strip()
                field_type = field.get('type', 'text')

                elements = item.select(field_selector)
                element = elements[0] if elements else None

                if field_type == 'text':
                    row[field_name] = extract_text(element)
                elif field_type == 'link':
                    row[field_name] = extract_link(element, base_url)
                elif field_type == 'image':
                    row[field_name] = extract_image(element, base_url)
                elif field_type == 'price':
                    row[field_name] = extract_price(element)
                elif field_type == 'html':
                    row[field_name] = str(element) if element else ''
                elif field_type == 'attribute':
                    attr = field.get('attribute', '').strip() or 'href'
                    row[field_name] = urljoin(base_url, element.get(attr, '')) if element else ''
                else:
                    row[field_name] = extract_text(element)
            row['_source_url'] = base_url
            return row

        def find_next_url(page_soup, base_url):
            selectors = [pagination_selector] if pagination_selector else []
            selectors.extend([s for s in common_next_selectors if s not in selectors])

            for selector in selectors:
                if not selector:
                    continue
                try:
                    next_elem = page_soup.select_one(selector)
                except Exception:
                    continue
                if not next_elem:
                    continue
                href = next_elem.get('href')
                if not href:
                    child_link = next_elem.select_one('a[href]') if hasattr(next_elem, 'select_one') else None
                    href = child_link.get('href') if child_link else None
                if href:
                    return urljoin(base_url, href)

            # Text fallback for sites that do not use semantic classes.
            for link in page_soup.find_all('a', href=True):
                label = link.get_text(' ', strip=True).lower()
                rel = ' '.join(link.get('rel', [])).lower() if isinstance(link.get('rel'), list) else str(link.get('rel', '')).lower()
                aria = str(link.get('aria-label', '')).lower()
                if label in ('next', 'next ›', '›', '»') or 'next' in aria or 'next' in rel:
                    return urljoin(base_url, link.get('href'))
            return None

        for page_number in range(1, pages_limit + 1):
            if not current_url or current_url in visited_urls:
                break

            html = fetch_page(
                current_url,
                proxy_url=proxy_url,
                render=render_enabled,
                load_more_selector=load_more_selector if load_more_enabled else '',
                max_load_more_clicks=max_load_more_clicks if load_more_enabled else 0,
                wait_after_load=wait_after_load,
            )
            page_soup = BeautifulSoup(html, 'html.parser')
            visited_urls.append(current_url)
            scope = scoped_soup(page_soup)
            if scope is None:
                return jsonify({
                    'success': False,
                    'error': f'Container selector not found: {container_selector}'
                }), 400

            items = scope.select(item_selector)
            if page_number == 1 and not items:
                return jsonify({
                    'success': False,
                    'error': f'No items found with selector: {item_selector}'
                }), 400

            for local_index, item in enumerate(items):
                if len(results) >= max_rows:
                    break
                try:
                    results.append(extract_row(item, current_url))
                except Exception as e:
                    failed_items.append({
                        'page': page_number,
                        'index': local_index,
                        'error': str(e)
                    })

            if len(results) >= max_rows or page_number >= pages_limit:
                break

            next_url = find_next_url(page_soup, current_url)
            if not next_url or next_url in visited_urls:
                break
            current_url = next_url
            time.sleep(0.35)

        return jsonify({
            'success': True,
            'data': results,
            'failed': failed_items,
            'visitedUrls': visited_urls,
            'stats': {
                'total_items': len(results),
                'failed_items': len(failed_items),
                'fields': len(fields),
                'pages_requested': pages_limit,
                'pages_scraped': len(visited_urls),
                'auto_detected': bool(auto_config_used),
                'item_selector': item_selector,
                'pagination_selector': pagination_selector,
                'proxy_enabled': proxy_enabled,
                'browser_rendering': False,
                'load_more_enabled': False,
                'load_more_selector': '',
                'max_load_more_clicks': 0,
                'edition': 'NO-CHROMIUM'
            },
            'autoConfig': auto_config_used,
            'fieldsUsed': fields
        })

    except Exception as e:
        return jsonify({'success': False, 'error': friendly_error(e)}), 500



def safe_sheet_title(name, fallback='Site'):
    name = normalize_text(name or fallback)
    name = re.sub(r'[\\/*?:\[\]]+', '_', name).strip() or fallback
    return name[:31]


def unique_sheet_title(wb, title):
    base = safe_sheet_title(title)
    candidate = base
    n = 2
    while candidate in wb.sheetnames:
        suffix = f'_{n}'
        candidate = base[:31-len(suffix)] + suffix
        n += 1
    return candidate


def dataset_name_from_url(url, index=1):
    parsed = urlparse(url or '')
    host = parsed.netloc or f'Site_{index}'
    path = parsed.path.strip('/').replace('/', '_')
    name = host if not path else f'{host}_{path[:18]}'
    return safe_sheet_title(name, f'Site_{index}')


@app.route('/api/multi-scrape', methods=['POST'])
def multi_scrape():
    """Scrape multiple websites in one run. Each website may have its own selectors, fields, and limits."""
    try:
        data = request.json or {}
        raw_websites = data.get('websites') or []

        websites = []
        if isinstance(raw_websites, list) and raw_websites:
            for index, item in enumerate(raw_websites, 1):
                if not isinstance(item, dict):
                    continue
                site_url = clean_url(item.get('url'))
                if not site_url:
                    continue
                websites.append({
                    'name': item.get('name') or dataset_name_from_url(site_url, index),
                    'url': site_url,
                    'containerSelector': item.get('containerSelector', ''),
                    'itemSelector': item.get('itemSelector', ''),
                    'fields': item.get('fields', []),
                    'maxRows': int(item.get('maxRows') or data.get('maxRows') or 100),
                    'maxPages': int(item.get('maxPages') or data.get('maxPages') or 1),
                    'paginationSelector': item.get('paginationSelector', '')
                })
        else:
            urls_text = data.get('urls', '') or data.get('multiUrls', '') or ''
            urls = [clean_url(line.strip()) for line in urls_text.splitlines() if line.strip()]
            for index, site_url in enumerate([u for u in urls if u], 1):
                websites.append({
                    'name': dataset_name_from_url(site_url, index),
                    'url': site_url,
                    'containerSelector': data.get('containerSelector', ''),
                    'itemSelector': data.get('itemSelector', ''),
                    'fields': data.get('fields', []),
                    'maxRows': int(data.get('maxRows') or 100),
                    'maxPages': int(data.get('maxPages') or 1),
                    'paginationSelector': data.get('paginationSelector', '')
                })

        if not websites:
            return jsonify({'success': False, 'error': 'At least one URL is required'}), 400

        for site in websites:
            ok, message = validate_url_value(site['url'])
            if not ok:
                return jsonify({'success': False, 'error': f"{message} Invalid entry: {site['url']}"}), 400
            ok, message = validate_scrape_options(site.get('maxRows', 100), site.get('maxPages', 1), 0)
            if not ok:
                return jsonify({'success': False, 'error': f"{site.get('name')}: {message}"}), 400

        datasets = []
        errors = []
        total_rows = 0
        combined_fields = {}

        with app.test_client() as client:
            for index, site in enumerate(websites, 1):
                payload = dict(data)
                payload.update({
                    'url': site['url'],
                    'containerSelector': site.get('containerSelector', ''),
                    'itemSelector': site.get('itemSelector', ''),
                    'fields': site.get('fields', []),
                    'maxRows': site.get('maxRows', 100),
                    'maxPages': site.get('maxPages', 1),
                    'paginationSelector': site.get('paginationSelector', ''),
                    'renderEnabled': False,
                    'loadMoreEnabled': False
                })
                payload.pop('urls', None)
                payload.pop('multiUrls', None)
                payload.pop('websites', None)

                response = client.post('/api/scrape', json=payload)
                result = response.get_json(silent=True) or {}
                if response.status_code == 200 and result.get('success'):
                    name = safe_sheet_title(site.get('name') or dataset_name_from_url(site['url'], index), f'Site_{index}')
                    fields_used = result.get('fieldsUsed') or site.get('fields') or []
                    rows = result.get('data', [])
                    for row in rows:
                        row['_website'] = name
                    for field in fields_used:
                        fname = field.get('name')
                        if fname:
                            combined_fields[fname] = field
                    combined_fields['_website'] = {'name': '_website', 'type': 'text'}
                    datasets.append({
                        'name': name,
                        'url': site['url'],
                        'data': rows,
                        'fields': fields_used + [{'name': '_website', 'type': 'text'}],
                        'stats': result.get('stats', {}),
                        'failed': result.get('failed', []),
                        'visitedUrls': result.get('visitedUrls', [])
                    })
                    total_rows += len(rows)
                else:
                    errors.append({'url': site['url'], 'name': site.get('name') or f'Site {index}', 'error': result.get('error', 'Scraping failed')})

        if not datasets:
            return jsonify({'success': False, 'error': 'All websites failed', 'errors': errors}), 400

        combined_rows = []
        for dataset in datasets:
            combined_rows.extend(dataset.get('data', []))
        if len(datasets) > 1:
            datasets.insert(0, {
                'name': 'Combined',
                'url': '',
                'data': combined_rows,
                'fields': list(combined_fields.values()),
                'stats': {'total_items': len(combined_rows)},
                'failed': [],
                'visitedUrls': []
            })

        return jsonify({
            'success': True,
            'datasets': datasets,
            'errors': errors,
            'stats': {
                'sites_requested': len(websites),
                'sites_scraped': len(datasets) - (1 if len(datasets) > 1 else 0),
                'sites_failed': len(errors),
                'total_items': total_rows
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': friendly_error(e)}), 500


@app.route('/api/link-scrape', methods=['POST'])
def link_scrape():
    """Execute link scraping"""
    try:
        data = request.json
        urls_text = data.get('urls', '').strip()
        filter_type = data.get('filterType', 'all')  # all, internal, external, same_domain
        remove_duplicates = data.get('removeDuplicates', True)
        max_links = data.get('maxLinks', 1000)
        
        if not urls_text:
            return jsonify({'success': False, 'error': 'At least one URL is required'}), 400
        
        # Parse URLs (one per line)
        urls = [clean_url(line.strip()) for line in urls_text.split('\n') if line.strip()]
        
        if not urls:
            return jsonify({'success': False, 'error': 'No valid URLs provided'}), 400
        
        all_links = []
        seen = set()
        
        for source_url in urls:
            try:
                html = fetch_page(source_url)
                soup = BeautifulSoup(html, 'html.parser')
                
                # Get base domain for filtering
                base_domain = urlparse(source_url).netloc
                
                # Extract all links
                for link_elem in soup.find_all('a', href=True):
                    href = link_elem.get('href', '')
                    if not href or href.startswith('#'):
                        continue
                    
                    # Convert to absolute URL
                    absolute_url = urljoin(source_url, href)
                    anchor_text = extract_text(link_elem)
                    
                    # Determine link type
                    link_type = get_link_type(absolute_url, base_domain)
                    
                    # Apply filters
                    if filter_type == 'internal' and link_type != 'internal':
                        continue
                    if filter_type == 'external' and link_type != 'external':
                        continue
                    if filter_type == 'same_domain':
                        link_domain = urlparse(absolute_url).netloc
                        if link_domain != base_domain:
                            continue
                    
                    # Remove duplicates
                    if remove_duplicates:
                        if absolute_url in seen:
                            continue
                        seen.add(absolute_url)
                    
                    all_links.append({
                        'source_url': source_url,
                        'found_url': absolute_url,
                        'anchor_text': anchor_text or '(no text)',
                        'link_type': link_type,
                        'scraped_at': datetime.now().isoformat()
                    })
                    
                    if len(all_links) >= max_links:
                        break
                
                if len(all_links) >= max_links:
                    break
                    
            except Exception as e:
                continue
        
        return jsonify({
            'success': True,
            'data': all_links,
            'stats': {
                'total_links': len(all_links),
                'sources': len(urls)
            }
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/test-selector', methods=['POST'])
def test_selector():
    """Test a CSS selector on a page"""
    try:
        data = request.json
        url = clean_url(data.get('url'))
        selector = data.get('selector', '').strip()
        
        if not url:
            return jsonify({'success': False, 'error': 'URL is required'}), 400
        
        if not selector:
            return jsonify({'success': False, 'error': 'Selector is required'}), 400
        
        # Fetch page
        html = fetch_page(url)
        soup = BeautifulSoup(html, 'html.parser')
        
        # Test selector
        elements = soup.select(selector)
        
        if not elements:
            return jsonify({
                'success': True,
                'count': 0,
                'message': 'No elements found with this selector',
                'samples': []
            })
        
        # Get first 5 samples
        samples = []
        for elem in elements[:5]:
            text = extract_text(elem)
            samples.append(text[:200] if text else '(empty)')
        
        return jsonify({
            'success': True,
            'count': len(elements),
            'message': f'Found {len(elements)} element(s)',
            'samples': samples
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/analyze-page', methods=['POST'])
def analyze_page():
    """Analyze page structure and suggest selectors"""
    try:
        data = request.json
        url = clean_url(data.get('url'))
        
        if not url:
            return jsonify({'success': False, 'error': 'URL is required'}), 400
        
        # Fetch page
        html = fetch_page(url)
        soup = BeautifulSoup(html, 'html.parser')
        
        # Collect statistics
        all_tags = [elem.name for elem in soup.find_all()]
        tag_counts = Counter(all_tags).most_common(15)
        
        # Collect classes
        all_classes = []
        for elem in soup.find_all(class_=True):
            classes = elem.get('class', [])
            if isinstance(classes, list):
                all_classes.extend(classes)
            else:
                all_classes.append(classes)
        
        class_counts = Counter(all_classes).most_common(20)
        
        # Suggest container selectors (common container patterns)
        container_suggestions = []
        for selector in ['.container', '.main', '.content', '#content', 'main', 'article', '.wrapper']:
            elems = soup.select(selector)
            if elems:
                sample_text = extract_text(elems[0])[:100]
                container_suggestions.append({
                    'selector': selector,
                    'count': len(elems),
                    'sample': sample_text
                })
        
        # Suggest item selectors (repeated patterns)
        item_suggestions = []
        for class_name, count in class_counts[:10]:
            if count >= 3:  # Repeated at least 3 times
                selector = f'.{class_name}'
                elems = soup.select(selector)
                if elems:
                    sample_text = extract_text(elems[0])[:100]
                    item_suggestions.append({
                        'selector': selector,
                        'count': count,
                        'sample': sample_text
                    })
        
        # Add common item patterns
        for selector in ['.item', '.product', '.post', '.card', 'article', '.entry']:
            elems = soup.select(selector)
            if elems and len(elems) >= 2:
                sample_text = extract_text(elems[0])[:100]
                if not any(s['selector'] == selector for s in item_suggestions):
                    item_suggestions.append({
                        'selector': selector,
                        'count': len(elems),
                        'sample': sample_text
                    })
        
        return jsonify({
            'success': True,
            'tags': [{'tag': tag, 'count': count} for tag, count in tag_counts],
            'classes': [{'class': cls, 'count': count} for cls, count in class_counts],
            'containerSuggestions': container_suggestions[:10],
            'itemSuggestions': item_suggestions[:10]
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/export/<format>', methods=['POST'])
def export_data(format):
    """Export data in various formats"""
    try:
        if format not in ('csv', 'xlsx', 'html'):
            return jsonify({'success': False, 'error': 'Invalid export format. Use Excel, CSV, or Web Page only.'}), 400
        data = request.json
        rows = data.get('rows', [])
        fields = data.get('fields', [])
        datasets = data.get('datasets', []) or []
        if datasets:
            datasets = [d for d in datasets if d.get('data')]
        
        if not rows and not datasets:
            return jsonify({'success': False, 'error': 'No data to export'}), 400
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        os.makedirs('output', exist_ok=True)

        if datasets and format == 'xlsx':
            filename = f'basira_multi_site_{timestamp}.xlsx'
            filepath = os.path.join('output', filename)
            wb = Workbook()
            wb.remove(wb.active)
            header_fill = PatternFill(start_color='0EA5E9', end_color='0EA5E9', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            for idx, dataset in enumerate(datasets, 1):
                ws = wb.create_sheet(unique_sheet_title(wb, dataset.get('name') or f'Site_{idx}'))
                dataset_fields = dataset.get('fields') or []
                headers = [f.get('name') for f in dataset_fields if f.get('name')]
                if not headers and dataset.get('data'):
                    headers = list(dataset['data'][0].keys())
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='left')
                for row_idx, row in enumerate(dataset.get('data', []), 2):
                    for col_idx, header in enumerate(headers, 1):
                        ws.cell(row=row_idx, column=col_idx, value=row.get(header, ''))
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            max_length = max(max_length, len(str(cell.value or '')))
                        except Exception:
                            pass
                    ws.column_dimensions[column].width = min(max_length + 2, 55)
            wb.save(filepath)
            return send_file(filepath, as_attachment=True, download_name=filename)

        if datasets and format == 'csv':
            import csv, zipfile
            from io import StringIO
            filename = f'basira_multi_site_csv_{timestamp}.zip'
            filepath = os.path.join('output', filename)
            with zipfile.ZipFile(filepath, 'w', zipfile.ZIP_DEFLATED) as zf:
                for idx, dataset in enumerate(datasets, 1):
                    dataset_fields = dataset.get('fields') or []
                    headers = [f.get('name') for f in dataset_fields if f.get('name')]
                    if not headers and dataset.get('data'):
                        headers = list(dataset['data'][0].keys())
                    output = StringIO()
                    writer = csv.DictWriter(output, fieldnames=headers, extrasaction='ignore')
                    writer.writeheader()
                    writer.writerows(dataset.get('data', []))
                    zf.writestr((safe_sheet_title(dataset.get('name') or f'Site_{idx}') + '.csv'), output.getvalue().encode('utf-8-sig'))
            return send_file(filepath, as_attachment=True, download_name=filename)

        if datasets and format == 'html':
            filename = f'basira_multi_site_{timestamp}.html'
            filepath = os.path.join('output', filename)
            def esc(v):
                return html_lib.escape(normalize_text(v))
            sections = []
            for idx, dataset in enumerate(datasets, 1):
                dataset_fields = dataset.get('fields') or []
                headers = [f.get('name') for f in dataset_fields if f.get('name')]
                if not headers and dataset.get('data'):
                    headers = list(dataset['data'][0].keys())
                table_head = ''.join(f'<th>{esc(h)}</th>' for h in headers)
                table_rows = []
                for row in dataset.get('data', []):
                    cells = ''.join(f'<td>{esc(row.get(h, ""))}</td>' for h in headers)
                    table_rows.append(f'<tr>{cells}</tr>')
                sections.append(f'<section><h2>{esc(dataset.get("name") or f"Site {idx}")}</h2><div class="meta">URL: {esc(dataset.get("url", ""))} | Rows: {len(dataset.get("data", []))}</div><div class="table-wrap"><table><thead><tr>{table_head}</tr></thead><tbody>{"".join(table_rows)}</tbody></table></div></section>')
            html_doc = f'''<!doctype html><html lang="en"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1"><title>Basira Multi Site Results</title><style>body{{font-family:Arial,system-ui,sans-serif;background:#f8fafc;color:#0f172a;margin:24px}}h1{{font-size:24px}}h2{{margin-top:28px;color:#0369a1}}.table-wrap{{overflow:auto;border:1px solid #cbd5e1;border-radius:14px;background:white}}table{{border-collapse:collapse;width:100%;font-size:14px}}th{{background:#dbeafe;color:#0369a1;text-align:left;padding:10px;border-bottom:1px solid #cbd5e1;position:sticky;top:0}}td{{padding:10px;border-bottom:1px solid #e2e8f0;vertical-align:top}}tr:hover td{{background:#f1f5f9}}.meta{{color:#64748b;margin-bottom:14px}}</style></head><body><h1>Basira Multi Site Results</h1><div class="meta">Websites: {len(datasets)} | Exported at: {datetime.now().isoformat(timespec='seconds')}</div>{''.join(sections)}</body></html>'''
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(html_doc)
            return send_file(filepath, as_attachment=True, download_name=filename)
        
        if format == 'csv':
            import csv
            from io import StringIO
            
            output = StringIO()
            writer = csv.DictWriter(output, fieldnames=[f['name'] for f in fields])
            writer.writeheader()
            writer.writerows(rows)
            
            filename = f'basira_scrape_{timestamp}.csv'
            filepath = os.path.join('output', filename)
            
            with open(filepath, 'w', encoding='utf-8', newline='') as f:
                f.write(output.getvalue())
            
            return send_file(filepath, as_attachment=True, download_name=filename)
        
        elif format == 'html':
            filename = f'basira_scrape_{timestamp}.html'
            filepath = os.path.join('output', filename)
            headers = [f['name'] for f in fields]
            def esc(v):
                return html_lib.escape(normalize_text(v))
            table_head = ''.join(f'<th>{esc(h)}</th>' for h in headers)
            table_rows = []
            for row in rows:
                cells = ''.join(f'<td>{esc(row.get(h, ""))}</td>' for h in headers)
                table_rows.append(f'<tr>{cells}</tr>')
            html_doc = f"""<!doctype html>
<html lang=\"en\">
<head>
<meta charset=\"utf-8\">
<meta name=\"viewport\" content=\"width=device-width, initial-scale=1\">
<title>Basira Scrape Results</title>
<style>
body{{font-family:Arial,system-ui,sans-serif;background:#f8fafc;color:#0f172a;margin:24px}}
h1{{font-size:22px}}
.table-wrap{{overflow:auto;border:1px solid #cbd5e1;border-radius:14px;background:white}}
table{{border-collapse:collapse;width:100%;font-size:14px}}
th{{background:#dbeafe;color:#0369a1;text-align:left;padding:10px;border-bottom:1px solid #cbd5e1;position:sticky;top:0}}
td{{padding:10px;border-bottom:1px solid #e2e8f0;vertical-align:top}}
tr:hover td{{background:#f1f5f9}}
.meta{{color:#64748b;margin-bottom:14px}}
</style>
</head>
<body>
<h1>Basira Scrape Results</h1>
<div class=\"meta\">Rows: {len(rows)} | Exported at: {datetime.now().isoformat(timespec='seconds')}</div>
<div class=\"table-wrap\"><table><thead><tr>{table_head}</tr></thead><tbody>{''.join(table_rows)}</tbody></table></div>
</body>
</html>"""
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(html_doc)
            return send_file(filepath, as_attachment=True, download_name=filename)
        
        elif format == 'xlsx':
            filename = f'basira_scrape_{timestamp}.xlsx'
            filepath = os.path.join('output', filename)
            
            wb = Workbook()
            ws = wb.active
            ws.title = 'Scrape Results'
            
            # Header style
            header_fill = PatternFill(start_color='0EA5E9', end_color='0EA5E9', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            # Write headers
            headers = [f['name'] for f in fields]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='left')
            
            # Write data
            for row_idx, row in enumerate(rows, 2):
                for col_idx, field in enumerate(fields, 1):
                    value = row.get(field['name'], '')
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Auto-size columns
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            wb.save(filepath)
            
            return send_file(filepath, as_attachment=True, download_name=filename)
        
        else:
            return jsonify({'success': False, 'error': 'Invalid format'}), 400
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# ═══════════════════════════════════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════════════════════════════════

if __name__ == '__main__':
    # Check if port is in use
    if is_port_in_use(BASIRA_PORT):
        print(f"ERROR: Port {BASIRA_PORT} is already in use.")
        print("Please close the application using this port or set another BASIRA_PORT.")
        sys.exit(1)

    print("Basira Web Scraping Local")
    print(f"Running at: http://{BASIRA_HOST}:{BASIRA_PORT}")
    print("Press Ctrl+C to stop.")

    app.run(host=BASIRA_HOST, port=BASIRA_PORT, debug=False)
